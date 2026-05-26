# -*- coding: utf-8 -*-
"""Génère ``test_data/matrice_tests.xlsx`` — 189 tests de qualification PhotoOrganizer.

Source de vérité unique pour la matrice : éditer ce fichier puis relancer.
Sortie : feuilles Tests, Synthèse, Exigences, Légende.

Mise à jour 2026-05-18 (v2.3-dev) :
  - 15 tests ajoutés (T-151 à T-165) pour la refonte v2.3 :
      • Onglets internes du panneau Organisation (4 onglets)
      • Bouton 💡 Exemples de marques + panneau intégré
      • Remplacement des fenêtres modales par des panneaux intégrés
  - Mises à jour de T-001/T-007/T-008/T-028/T-140 (IHM refondue).

Mise à jour 2026-05-19 :
  - 10 tests ajoutés (T-166 à T-175) pour le bouton 💡 Exemples de filtres
    de l'onglet Filtrer (mots-clés, extensions, dimensions, orientation,
    note — valeurs standards qui aident l'utilisateur à éviter les erreurs
    de remplissage des champs de pré-filtre).

Mise à jour 2026-05-19 (II) — Quarantaine réversible :
  - 14 tests ajoutés (T-176 à T-189) pour le système de quarantaine
    interne qui rend le mode TRASH des doublons réversible. Avant :
    send2trash était un aller simple. Après : déplacement vers une
    quarantaine interne récupérable via « Annuler dernière/tout », puis
    bouton « 🗑 Vider quarantaine » pour passer en corbeille système.

Usage :
    python test_data/scripts/generate_matrix.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Définition des 150 tests
# ---------------------------------------------------------------------------
COLUMNS = [
    "ID_Test", "Catégorie", "Exigence liée", "Fonctionnalité",
    "Description", "Pré-requis", "Données entrée",
    "Résultat attendu", "Résultat obtenu", "Statut",
    "Sévérité", "Testeur", "Date", "Commentaires",
]

# Sévérités : Bloquant / Majeur / Mineur / Cosmétique
# Statut (rempli à l'exécution) : OK / NOK / NA / À faire


def T(idx, categorie, req, fonctionnalite, description,
      prerequis, entree, attendu, severite):
    """Helper de construction d'une ligne test."""
    return [
        f"T-{idx:03d}", categorie, req, fonctionnalite, description,
        prerequis, entree, attendu, "", "À faire",
        severite, "", "", "",
    ]


tests = []

# =============================================================================
# IHM (T-001 à T-030)
# =============================================================================
tests += [
    T(1, "IHM", "REQ-IHM-01", "Démarrage de l'application", "Démarrer l'application : 4 onglets principaux (📁/🔍/📜/⚙️) + barre d'état « Prêt »", "Exécutable compilé", "—", "Bandeau de 4 onglets, fenêtre 1200×800, panneau Organisation affiche lui-même 4 onglets internes (v2.3)", "Bloquant"),
    T(2, "IHM", "REQ-IHM-02", "Navigation onglets", "Ctrl+1..4 navigue entre onglets, F1 ouvre À propos", "T-001 OK", "—", "Onglet actif change ; modale About s'ouvre", "Majeur"),
    T(3, "IHM", "REQ-IHM-03", "Bouton thème ☀️/🌙", "Clic bascule entre clair/sombre", "T-001 OK", "—", "Frames et boutons changent de couleur", "Mineur"),
    T(4, "IHM", "REQ-IHM-04", "minsize fenêtre", "Réduire la fenêtre à 800×550", "T-001 OK", "—", "Tous les boutons restent atteignables, scroll central activé si besoin", "Majeur"),
    T(5, "IHM", "REQ-IHM-05", "Tooltips bouton Organiser", "Survoler 600 ms le bouton 🚀 Organiser", "T-001 OK", "—", "Info-bulle FR apparaît avec description action+effet", "Mineur"),
    T(6, "IHM", "REQ-IHM-06", "Sticky bottom Organisation", "Fenêtre 1200×800 + déplier Avancé", "T-001 OK", "—", "Boutons d'action restent visibles, scroll central scrollable", "Majeur"),
    T(7, "IHM", "REQ-IHM-07", "Onglet Filtrer accessible", "Cliquer l'onglet 🔍 Filtrer dans le bandeau d'onglets interne du panneau Organisation", "T-001 OK", "—", "Contenu Filtres visible, filtres date/taille/note/mots-clés accessibles (refonte v2.3 — remplace l'ancien bouton ▶ Avancé)", "Majeur"),
    T(8, "IHM", "REQ-IHM-08", "Onglet Renommer accessible", "Cliquer l'onglet 🏷 Renommer dans le bandeau d'onglets interne du panneau Organisation", "T-001 OK", "—", "Contenu Renommer visible, exemples cliquables présents (refonte v2.3 — remplace l'ancien bouton Renommage repliable)", "Mineur"),
    T(9, "IHM", "REQ-IHM-09", "Sidebar Doublons largeur ≥ 500", "Onglet Doublons : vérif options non tronquées", "T-001 OK", "—", "Tous champs/cases lisibles sans scroll horizontal", "Mineur"),
    T(10, "IHM", "REQ-IHM-10", "Compteur fichiers temps réel", "Sélectionner source nominal", "T-001 OK", "input_nominal/", "Label « 📂 N fichier(s) prêt(s) » mis à jour < 2 s", "Majeur"),
    T(11, "IHM", "REQ-IHM-11", "Sources multiples via +", "Cliquer + Source 2 fois sur 2 dossiers différents", "T-001 OK", "input_nominal/ + input_nominal_2/", "Source field contient 'p1;p2', compteur cumule", "Mineur"),
    T(12, "IHM", "REQ-IHM-12", "Drag & drop dossier", "Glisser un dossier depuis explorateur sur champ Source", "tkinterdnd2 OK", "input_nominal/", "Champ Source mis à jour avec chemin", "Mineur"),
    T(13, "IHM", "REQ-IHM-13", "Bouton ↗ Ouvrir destination", "Sélectionner dest, cliquer ↗", "Dest existe", "—", "Explorer Windows ouvre le dossier", "Mineur"),
    T(14, "IHM", "REQ-IHM-14", "Switch multicouche révèle ordre", "Cocher 2 critères puis activer switch", "T-001 OK", "—", "Sous-panneau Ordre apparaît avec 2 cards", "Majeur"),
    T(15, "IHM", "REQ-IHM-15", "Boutons ▲▼ réordonnent critères", "Multilayer ON, cliquer ▲ sur Camera", "T-014 OK", "—", "Camera passe au-dessus de Date dans la liste", "Mineur"),
    T(16, "IHM", "REQ-IHM-16", "Sous-options GPS conditionnelles", "Cocher Par localisation GPS", "T-001 OK", "—", "Géocodage + slider distance + label apparaissent", "Mineur"),
    T(17, "IHM", "REQ-IHM-17", "Slider distance max + boutons fins", "Cliquer ▶ 3 fois sur slider distance", "T-016 OK", "—", "Label passe de '1.0 km' à '1.3 km' (pas 100 m)", "Mineur"),
    T(18, "IHM", "REQ-IHM-18", "Exemples Renommage cliquables", "Déplier Renommage, cliquer un exemple", "T-001 OK", "—", "Champ Template renseigné, aperçu mis à jour", "Mineur"),
    T(19, "IHM", "REQ-IHM-19", "Bouton 🔄 Réinitialiser template", "Cliquer Réinitialiser après exemple", "T-018 OK", "—", "Template devient vide, aperçu '(aucun template)'", "Mineur"),
    T(20, "IHM", "REQ-IHM-20", "Aperçu live template", "Saisir '{date:%Y%m%d}_{counter:04d}'", "T-001 OK", "—", "Aperçu affiché : 'IMG_0001.jpg → 20260507_0042.jpg'", "Mineur"),
    T(21, "IHM", "REQ-IHM-21", "Preset save/load via menu", "Configurer + 💾 + saisir nom + Charger", "T-001 OK", "—", "Preset persisté dans config/presets/*.json, rechargé OK", "Majeur"),
    T(22, "IHM", "REQ-IHM-22", "Preset suppression confirm", "Sélectionner preset + 🗑 + confirmer", "T-021 OK", "—", "Preset supprimé, OptionMenu mis à jour", "Mineur"),
    T(23, "IHM", "REQ-IHM-23", "Mode Doublons recolore Exécuter", "Cocher DELETE puis TRASH puis MOVE puis DRY_RUN", "T-001 OK", "—", "Couleur bouton : rouge / orange / bleu / vert", "Mineur"),
    T(24, "IHM", "REQ-IHM-24", "Mode MOVE révèle dest", "Cocher MOVE dans Doublons", "T-001 OK", "—", "Champ Destination MOVE apparaît", "Majeur"),
    T(25, "IHM", "REQ-IHM-25", "Onglets Doublons Résultats/Détails", "Onglet Doublons : vérifier le bandeau d'onglets", "T-001 OK", "—", "Exactement 2 onglets : Résultats + Détails", "Mineur"),
    T(26, "IHM", "REQ-IHM-26", "Statistiques Historique compactes", "Onglet Historique vierge", "T-001 OK", "—", "Étiquette de statistiques affiche « Aucune opération » (compact)", "Cosmétique"),
    T(27, "IHM", "REQ-IHM-27", "Boutons Annuler désactivés si vide", "Onglet Historique vierge", "T-001 OK", "—", "↩️ Annuler dernière + ↩️ Annuler tout sont à l'état désactivé", "Majeur"),
    T(28, "IHM", "REQ-IHM-28", "Panneau intégré résultat + 📂 Ouvrir", "Lancer une organisation simple, attendre la fin", "T-001 OK", "input_nominal/", "Panneau intégré avec zone de texte + bouton « 📂 Ouvrir destination » — AUCUNE nouvelle fenêtre (refonte 2026-05-18)", "Mineur"),
    T(29, "IHM", "REQ-IHM-29", "Toast Windows fin organisation", "Cocher Notification système, organiser", "T-001 OK + plyer dispo", "input_nominal/", "Toast système non-modale visible 5s", "Mineur"),
    T(30, "IHM", "REQ-IHM-30", "Couleurs sémantiques cohérentes", "Inspecter visuellement les boutons d'action principale / danger / avertissement des 4 panneaux", "T-001 OK", "—", "Vert (principal) / Rouge (danger) / Orange (avertissement) uniformes", "Cosmétique"),
]

# =============================================================================
# Paramètres (T-031 à T-050)
# =============================================================================
tests += [
    T(31, "Paramètres", "REQ-PRM-01", "Sauvegarde du thème + persistance", "Onglet Paramètres : Sombre → Clair → Sauvegarder, fermer, relancer", "T-001 OK", "—", "Thème Clair restauré après redémarrage", "Majeur"),
    T(32, "Paramètres", "REQ-PRM-02", "Réinitialisation aux valeurs par défaut avec confirmation", "Cliquer Réinitialiser puis confirmer la boîte de dialogue", "T-001 OK", "—", "Toutes les valeurs reviennent aux valeurs par défaut de la configuration", "Majeur"),
    T(33, "Paramètres", "REQ-PRM-03", "Vider cache des métadonnées", "Cliquer 🗑 Vider cache + confirm", "T-001 OK + cache non vide", "—", "Stats cache passent à 0 entrées en mémoire et disque", "Mineur"),
    T(34, "Paramètres", "REQ-PRM-04", "Planification activation", "Switch ON + Heure '23:00' + Sauvegarder", "T-001 OK", "—", "Status « ⏰ Prochaine exécution : … 23:00 »", "Majeur"),
    T(35, "Paramètres", "REQ-PRM-05", "Heure invalide rejetée", "Saisir '9999:99' puis activer", "T-001 OK", "—", "Switch OFF ou statut « ⚠️ Heure invalide »", "Mineur"),
    T(36, "Paramètres", "REQ-PRM-06", "Action défaut Copy/Move", "Radio 'Déplacer' + Sauvegarder + relancer", "T-001 OK", "—", "Onglet Organisation Radio Move pré-cochée", "Mineur"),
    T(37, "Paramètres", "REQ-PRM-07", "Géocodage on/off", "Décocher Géocodage + Sauvegarder", "T-001 OK", "—", "GPSProcessor.geocoding_enabled = False au prochain run", "Mineur"),
    T(38, "Paramètres", "REQ-PRM-08", "Clé API PositionStack", "Saisir clé '1234abc' + Sauvegarder", "T-001 OK", "—", "Champ masqué (*), valeur persistée", "Mineur"),
    T(39, "Paramètres", "REQ-PRM-09", "Niveau log change", "Sélectionner DEBUG + Sauvegarder + relancer", "T-001 OK", "—", "Logs DEBUG visibles dans LOCALAPPDATA/PhotoOrganizer/logs/", "Mineur"),
    T(40, "Paramètres", "REQ-PRM-10", "Effacer dossiers récents", "Cliquer Effacer + confirm", "Source/Dest récents non vides", "—", "AppConfig.recent_sources/destinations vides", "Mineur"),
    T(41, "Paramètres", "REQ-PRM-11", "Window position/taille persistée", "Redimensionner fenêtre + fermer + relancer", "T-001 OK", "—", "Position/taille restaurées (window_width/height/x/y)", "Mineur"),
    T(42, "Paramètres", "REQ-PRM-12", "Récents source pré-rempli", "Utiliser source A puis fermer + relancer", "T-001 OK", "—", "Source A pré-renseignée dans Champ Source", "Mineur"),
    T(43, "Paramètres", "REQ-PRM-13", "Boîte de confirmation Sauvegarder", "Cliquer Sauvegarder", "Modifications en attente", "—", "Boîte de dialogue « Les paramètres ont été sauvegardés. »", "Cosmétique"),
    T(44, "Paramètres", "REQ-PRM-14", "Cache TTL respecté", "Forcer un scan, attendre TTL+1, relancer scan même fichier", "Cache enabled", "1 photo", "Re-lecture EXIF (cache miss après TTL)", "Mineur"),
    T(45, "Paramètres", "REQ-PRM-15", "max_cache_size_mb limite", "Saturer cache au-delà de la limite", "Cache enabled", "1000 photos", "Évictions LRU, taille reste sous la limite", "Mineur"),
    T(46, "Paramètres", "REQ-PRM-16", "Theme system auto-détection", "Sélectionner thème 'Système'", "T-001 OK", "—", "Suit le thème Windows (darkdetect)", "Mineur"),
    T(47, "Paramètres", "REQ-PRM-17", "Schedule trigger callback", "Activer schedule à HH:MM imminent (1 min)", "T-001 OK + source/dest config", "—", "_organize_files() déclenché automatiquement", "Majeur"),
    T(48, "Paramètres", "REQ-PRM-18", "Schedule idempotence", "Activer schedule, attendre 2 mn même HH:MM", "T-047 OK", "—", "Un seul trigger (last_run_date check)", "Majeur"),
    T(49, "Paramètres", "REQ-PRM-19", "Recursive défaut switch", "Décocher recursive défaut + Sauvegarder + relancer", "T-001 OK", "—", "Case 'Parcourir sous-dossiers' décochée par défaut", "Mineur"),
    T(50, "Paramètres", "REQ-PRM-20", "Récupération après configuration corrompue", "Modifier manuellement config.json (clé invalide)", "T-001 OK", "Configuration corrompue", "L'application démarre avec les valeurs par défaut + avertissement dans le journal", "Majeur"),
]

# =============================================================================
# Entrées (T-051 à T-065)
# =============================================================================
tests += [
    T(51, "Entrées", "REQ-ENT-01", "Cas nominal — 50 photos", "Lancer l'organisation sur le dossier nominal", "T-001 OK", "input_nominal/ (50 .jpg)", "Tous traités : 50/50 fichiers traités, 0 erreur", "Bloquant"),
    T(52, "Entrées", "REQ-ENT-02", "input_vide", "Lancer organisation sur dossier vide", "T-001 OK", "input_vide/ (0 fichier)", "Message 'Aucun fichier trouvé', 0 op", "Majeur"),
    T(53, "Entrées", "REQ-ENT-03", "input_volumineux 1000 photos", "Lancer organisation sur 1000 photos", "T-001 OK", "input_volumineux/ (1000 .jpg)", "Pas de freeze IHM, progress 0→100%, < 60 s", "Majeur"),
    T(54, "Entrées", "REQ-ENT-04", "input_mauvais_format", "Source contient .txt, .pdf, .docx", "T-001 OK", "input_mauvais_format/", "Filtrés par extensions, 0 photo traitée", "Mineur"),
    T(55, "Entrées", "REQ-ENT-05", "input_caracteres_speciaux", "Noms avec accents, emojis, μ, Ω, ±", "T-001 OK", "input_caracteres_speciaux/", "Sanitization OK, pas de crash, noms valides", "Majeur"),
    T(56, "Entrées", "REQ-ENT-06", "Fichier JPG corrompu (tronqué)", "Photo JPG dont les 50 % derniers octets manquent", "T-001 OK", "input_corrompu/", "Ignoré avec avertissement dans le journal, 1 erreur, autres fichiers traités", "Majeur"),
    T(57, "Entrées", "REQ-ENT-07", "input_gps_piexif rationnels", "Photos avec EXIF GPS encodé piexif ((48,1)(51,1)(24,1))", "T-001 OK + GPS coché", "input_gps_piexif/", "Coordonnées correctement extraites, organisées par lieu", "Majeur"),
    T(58, "Entrées", "REQ-ENT-08", "input_pairs RAW+JPEG", "10 paires .CR2 + .JPG même stem", "T-001 OK + keep_pairs ON", "input_pairs/", "Pairs co-localisées dans même dossier dest", "Mineur"),
    T(59, "Entrées", "REQ-ENT-09", "input_bursts 5 photos < 3s", "5 photos timestamps espacés de 1s", "T-001 OK + detect_bursts ON", "input_bursts/", "Sous-dossier Burst_01/ avec 5 fichiers", "Mineur"),
    T(60, "Entrées", "REQ-ENT-10", "input_pas_exif", "3 photos sans aucune métadonnée EXIF", "T-001 OK + date critère", "input_pas_exif/", "Tous classés dans dossier 'Sans date'", "Majeur"),
    T(61, "Entrées", "REQ-ENT-11", "input_doublons 10 paires", "20 photos dont 10 doublons exacts", "T-001 OK", "input_doublons/", "DRY_RUN détecte 10 groupes de 2", "Bloquant"),
    T(62, "Entrées", "REQ-ENT-12", "input HEIC iPhone", "Photo HEIC avec EXIF", "pillow_heif installé", "input_heic/", "Lue, organisée, métadonnées extraites", "Majeur"),
    T(63, "Entrées", "REQ-ENT-13", "input vidéo MP4 avec date", "1 vidéo MP4 avec CreationTime EXIF", "include_videos ON", "input_video/", "Organisée par date comme une photo", "Mineur"),
    T(64, "Entrées", "REQ-ENT-14", "input_limite_haute 100 Mo", "1 fichier RAW de 100 Mo", "T-001 OK", "input_limite_haute/", "Traité sans timeout, log temps de copie", "Mineur"),
    T(65, "Entrées", "REQ-ENT-15", "input_limite_basse 1 Ko", "1 photo de < 1 Ko (vignette)", "filter_size_min OFF", "input_limite_basse/", "Traitée comme une photo normale", "Cosmétique"),
]

# =============================================================================
# Sorties (T-066 à T-080)
# =============================================================================
tests += [
    T(66, "Sorties", "REQ-SOR-01", "Arbo year/month/day", "Format date par défaut sur nominal", "T-001 OK", "input_nominal/", "Structure dest/YYYY/MM/YYYY_MM_DD/*.jpg", "Bloquant"),
    T(67, "Sorties", "REQ-SOR-02", "Arbo year/month", "Format year/month", "T-066 OK", "input_nominal/", "Structure dest/YYYY/YYYY_MM/*.jpg", "Mineur"),
    T(68, "Sorties", "REQ-SOR-03", "Arbo multilayer date+camera", "Date+Camera multilayer", "T-014 OK", "input_nominal/", "dest/YYYY/MM/.../CameraModel/*.jpg", "Majeur"),
    T(69, "Sorties", "REQ-SOR-04", "Index CSV généré", "Cocher Index CSV", "T-001 OK", "input_nominal/", "dest/_photoorganizer_index_*.csv, colonnes valides", "Majeur"),
    T(70, "Sorties", "REQ-SOR-05", "Index JSON valide", "Cocher Index JSON", "T-001 OK", "input_nominal/", "Fichier JSON valide ; nombre d'enregistrements == nombre de fichiers traités", "Majeur"),
    T(71, "Sorties", "REQ-SOR-06", "Rapport Doublons TXT", "Mode DRY_RUN + case TXT", "T-001 OK", "input_doublons/", "duplicates_report_*.txt généré et lisible", "Mineur"),
    T(72, "Sorties", "REQ-SOR-07", "Rapport Doublons CSV", "Mode DRY_RUN + case CSV", "T-001 OK", "input_doublons/", "CSV exploitable Excel", "Mineur"),
    T(73, "Sorties", "REQ-SOR-08", "Rapport Doublons JSON", "Mode DRY_RUN + case JSON", "T-001 OK", "input_doublons/", "JSON structuré (groups, decisions, …)", "Mineur"),
    T(74, "Sorties", "REQ-SOR-09", "Renommage template {date}_{counter}", "Template '{date:%Y%m%d}_{counter:04d}'", "T-001 OK", "input_nominal/", "Fichiers : 20260507_0001.jpg, 20260507_0002.jpg, …", "Majeur"),
    T(75, "Sorties", "REQ-SOR-10", "Renommage tokens manquants fallback", "Template avec {camera} sur photo sans EXIF", "T-001 OK", "input_pas_exif/", "Fallback 'unknown' au lieu de crash", "Mineur"),
    T(76, "Sorties", "REQ-SOR-11", "Burst sous-dossier Burst_NN", "detect_bursts ON, 5 photos < 3s", "T-001 OK", "input_bursts/", "Sous-dossier Burst_01/ contient les 5", "Mineur"),
    T(77, "Sorties", "REQ-SOR-12", "Cleanup post-MOVE empty source dirs", "MOVE + cleanup ON", "T-001 OK", "input_nominal/", "Sous-dossiers source vides supprimés à la fin", "Mineur"),
    T(78, "Sorties", "REQ-SOR-13", "Index incrémental persisté", "Mode incrémental ON, 2 runs", "T-001 OK", "input_nominal/", "dest/.photoorganizer_index.json créé et enrichi", "Majeur"),
    T(79, "Sorties", "REQ-SOR-14", "Sans date → dossier 'Sans date'", "Date critère + photo sans EXIF", "T-001 OK", "input_pas_exif/", "Photos placées dans dest/Sans date/", "Mineur"),
    T(80, "Sorties", "REQ-SOR-15", "GPS sans nom → Lat_x_Lon_y", "Géocodage off + GPS critère", "T-001 OK", "input_gps_piexif/", "Sous-dossiers Lat_48.85_Lon_2.35", "Mineur"),
]

# =============================================================================
# Cas limites (T-081 à T-105)
# =============================================================================
tests += [
    T(81, "Cas limites", "REQ-LIM-01", "Source = destination", "Choisir le même dossier en source et dest", "T-001 OK", "input_nominal/", "Refus avec message d'erreur explicite ou fonctionnement safe", "Majeur"),
    T(82, "Cas limites", "REQ-LIM-02", "Destination read-only", "Dest avec droits 'Lecture seule'", "T-001 OK", "input_nominal/", "Erreur OS catchée, message clair", "Majeur"),
    T(83, "Cas limites", "REQ-LIM-03", "Espace disque insuffisant", "Option « Valider espace disque » activée, destination pleine", "T-001 OK", "input_volumineux/", "Fenêtre d'avertissement avec choix Continuer/Annuler", "Majeur"),
    T(84, "Cas limites", "REQ-LIM-04", "Chemin avec espaces + UTF-8", "Source = 'D:\\Photos été 2024\\été #'", "T-001 OK", "input_nominal/", "Organisation OK, pas de crash", "Majeur"),
    T(85, "Cas limites", "REQ-LIM-05", "Fichier déjà existant — ignorer si identique", "Relancer l'organisation à l'identique avec « ignorer si identique » coché", "T-001 OK", "input_nominal/", "2e passage : 0 fichier traité, N fichiers ignorés", "Mineur"),
    T(86, "Cas limites", "REQ-LIM-06", "Trash re-scan exclu", "Mode TRASH puis re-list source", "T-001 OK", "input_doublons/", "$Recycle.Bin / .Trash-* exclus du scan", "Majeur"),
    T(87, "Cas limites", "REQ-LIM-07", "Multi-source déduplication", "source1;source1 (même chemin 2 fois)", "T-001 OK", "input_nominal/", "Fichiers dédupliqués via normcase+abspath", "Mineur"),
    T(88, "Cas limites", "REQ-LIM-08", "Template tokens sur photo sans date", "Template {date:%Y%m%d}_{original}", "T-001 OK", "input_pas_exif/", "Token date remplacé par 'unknown'", "Mineur"),
    T(89, "Cas limites", "REQ-LIM-09", "auto_rename collisions", "Re-run avec auto_rename ON", "T-001 OK", "input_nominal/", "Fichiers renommés _1, _2, _3 sans crash", "Majeur"),
    T(90, "Cas limites", "REQ-LIM-10", "Date filtre exclut tout", "date_min = 2099-01-01", "T-001 OK", "input_nominal/", "0 fichier éligible, message clair", "Mineur"),
    T(91, "Cas limites", "REQ-LIM-11", "Filtre rating exclut tout", "rating_min = 5, aucune photo notée", "T-001 OK", "input_nominal/", "0 fichier éligible", "Mineur"),
    T(92, "Cas limites", "REQ-LIM-12", "Filtre keywords avec OR", "keywords = 'vacances, mariage'", "T-001 OK", "input_keywords/", "Match si AU MOINS UN mot-clé présent", "Mineur"),
    T(93, "Cas limites", "REQ-LIM-13", "Seuil de rafale trop strict", "Seuil = 1 s, nombre minimum = 10 photos", "T-001 OK", "input_bursts/", "0 rafale détectée (groupes trop petits)", "Mineur"),
    T(94, "Cas limites", "REQ-LIM-14", "Multilayer avec ordre exotique", "Order ['location','camera','date']", "T-014 OK", "input_nominal_with_gps/", "Hiérarchie GPS/Camera/Date respectée", "Mineur"),
    T(95, "Cas limites", "REQ-LIM-15", "Aucun critère coché", "Tout décocher, lancer", "T-001 OK", "input_nominal/", "Erreur ou avertissement explicite, refus de l'opération ou copie à plat", "Majeur"),
    T(96, "Cas limites", "REQ-LIM-16", "Annulation pendant scan", "Cliquer Annuler dès le démarrage", "T-001 OK", "input_volumineux/", "Arrêt < 2 s, état stable, log info", "Majeur"),
    T(97, "Cas limites", "REQ-LIM-17", "Annulation à 50% organisation", "Annuler à mi-parcours", "T-001 OK", "input_volumineux/", "Arrêt < 2 s, fichiers déjà traités intacts", "Majeur"),
    T(98, "Cas limites", "REQ-LIM-18", "Preset YAML corrompu", "Charger un preset YAML invalide", "T-001 OK", "preset_corrompu.yaml", "Messagebox erreur, pas de crash", "Mineur"),
    T(99, "Cas limites", "REQ-LIM-19", "Doublons sur 0 fichier", "Rechercher dans dossier vide", "T-001 OK", "input_vide/", "Message 'Aucun doublon', pas de crash", "Mineur"),
    T(100, "Cas limites", "REQ-LIM-20", "Drag fichier (pas dossier)", "Glisser un .jpg sur champ Source", "tkinterdnd2 OK", "—", "Refus silencieux (path doit être dossier)", "Mineur"),
    T(101, "Cas limites", "REQ-LIM-21", "Hors-ligne géocodage", "Pas d'internet + Par location ON + géocodage ON", "Network down", "input_gps_piexif/", "Fallback Lat_x_Lon_y, pas de timeout long", "Majeur"),
    T(102, "Cas limites", "REQ-LIM-22", "Permissions destination", "Dest = C:\\Windows\\System32 (UAC requis)", "T-001 OK", "input_nominal/", "Erreur clair, pas de crash", "Majeur"),
    T(103, "Cas limites", "REQ-LIM-23", "Source supprimée pendant op", "Sup dossier source pendant scan", "T-001 OK", "input_nominal/", "Erreurs catchées par fichier, total cohérent", "Majeur"),
    T(104, "Cas limites", "REQ-LIM-24", "Schedule trigger app close", "Activer schedule, fermer app", "T-001 OK", "—", "Scheduler thread stoppe proprement", "Mineur"),
    T(105, "Cas limites", "REQ-LIM-25", "Cache disque saturé", "Cache TTL=0 + 10k scans", "T-001 OK", "input_volumineux/", "Évictions, pas de crash, taille respecte limite", "Mineur"),
]

# =============================================================================
# Performance (T-106 à T-115)
# =============================================================================
tests += [
    T(106, "Performance", "REQ-PRF-01", "Cold-start EXE release", "Lancer dist/PhotoOrganizer-2.0.0.exe depuis disque froid", "Build OK", "—", "< 5 s avant log 'Démarrage'", "Majeur"),
    T(107, "Performance", "REQ-PRF-02", "Cold-start EXE debug", "Lancer dist/PhotoOrganizer-2.0.0-debug.exe", "Build OK", "—", "< 8 s (logging verbose accepté)", "Mineur"),
    T(108, "Performance", "REQ-PRF-03", "Organisation 1000 photos < 60 s", "F01 sur input_volumineux", "T-001 OK", "input_volumineux/ (1000)", "Temps total < 60 s sur SSD", "Mineur"),
    T(109, "Performance", "REQ-PRF-04", "Scan 10k fichiers récursif", "list_files recursive sur 10k", "T-001 OK", "input_10k/", "< 5 s pour le scan (avant org)", "Mineur"),
    T(110, "Performance", "REQ-PRF-05", "Empreinte mode rapide sur 100 Mo", "Doublons en mode rapide activé, RAW de 100 Mo", "T-001 OK", "input_raw_100mb/", "< 1 s par fichier (empreinte début + fin uniquement)", "Mineur"),
    T(111, "Performance", "REQ-PRF-06", "UI réactive pendant scan", "Cliquer Annuler pendant scan en cours", "T-001 OK", "input_volumineux/", "Annulation propagée < 2 s", "Majeur"),
    T(112, "Performance", "REQ-PRF-07", "RAM repos < 200 MB", "App ouverte, aucun scan", "EXE release", "—", "Process Python sub-MEI : RAM < 200 MB", "Mineur"),
    T(113, "Performance", "REQ-PRF-08", "RAM pic 1000 photos", "Organisation 1000 photos", "T-001 OK", "input_volumineux/", "RAM pic < 500 MB", "Mineur"),
    T(114, "Performance", "REQ-PRF-09", "Cache hit rate", "Re-scan après cache plein", "Cache enabled", "input_nominal/", "Hit rate > 95% sur 2e scan", "Mineur"),
    T(115, "Performance", "REQ-PRF-10", "Benchmark list_files pytest", "pytest tests/perf/test_perf.py", "Tests OK", "1000 fichiers générés", "Médiane < 150 ms (cf. perf baseline)", "Mineur"),
]

# =============================================================================
# Robustesse (T-116 à T-135)
# =============================================================================
tests += [
    T(116, "Robustesse", "REQ-ROB-01", "Cliquer Annuler 100× sans op", "Spam le bouton Annuler", "T-001 OK", "—", "Pas de NPE, état stable, pas de fuite", "Majeur"),
    T(117, "Robustesse", "REQ-ROB-02", "Double-clic Organiser", "Double-clic rapide sur 🚀 Organiser", "T-001 OK", "input_nominal/", "Une seule exécution lancée", "Majeur"),
    T(118, "Robustesse", "REQ-ROB-03", "Fermeture pendant op", "Alt+F4 pendant scan", "T-001 OK", "input_volumineux/", "Cleanup propre, aucun fichier verrouillé", "Majeur"),
    T(119, "Robustesse", "REQ-ROB-04", "Source supprimée pendant op", "Supprimer source dans Explorer", "T-001 OK", "input_nominal/", "Erreurs catchées par fichier, total cohérent", "Majeur"),
    T(120, "Robustesse", "REQ-ROB-05", "Bouton Avancé — clics répétés 50×", "Cliquer le bouton 50 fois rapidement", "T-001 OK", "—", "Pas de fuite mémoire, état final stable", "Mineur"),
    T(121, "Robustesse", "REQ-ROB-06", "Hors-ligne géocodage", "Internet OFF + GPS + géocodage", "T-001 OK", "input_gps_piexif/", "Fallback Lat_x_Lon_y, log debug", "Majeur"),
    T(122, "Robustesse", "REQ-ROB-07", "Preset YAML corrompu", "Charger preset YAML invalide", "T-001 OK", "preset_corrompu.yaml", "Messagebox erreur, app continue", "Mineur"),
    T(123, "Robustesse", "REQ-ROB-08", "Cache SQLite corrompu", "Modifier manuellement le .db", "Cache enabled", "Cache corrompu", "Cache recréé silencieusement", "Mineur"),
    T(124, "Robustesse", "REQ-ROB-09", "Drag & drop fichier (pas dossier)", "Glisser .jpg sur Source", "tkinterdnd2 OK", "—", "Refus silencieux", "Mineur"),
    T(125, "Robustesse", "REQ-ROB-10", "Drag & drop multi-fichiers", "Glisser 2 dossiers d'un coup", "tkinterdnd2 OK", "input_nominal/ + input_pairs/", "Champ Source rempli avec p1;p2", "Mineur"),
    T(126, "Robustesse", "REQ-ROB-11", "Très long template", "Saisir template de 200 chars", "T-001 OK", "input_nominal/", "Template tronqué ou rendu sans crash, fichiers OK", "Mineur"),
    T(127, "Robustesse", "REQ-ROB-12", "Caractères interdits filename", "Photo dest contiendrait '/' ou ':'", "T-001 OK", "input_nominal/", "Sanitization auto (remplacement _)", "Majeur"),
    T(128, "Robustesse", "REQ-ROB-13", "Schedule heure passée", "Activer schedule à HH:MM déjà dépassé", "T-001 OK", "—", "Status 'Prochaine exécution : demain HH:MM'", "Mineur"),
    T(129, "Robustesse", "REQ-ROB-14", "Schedule modifié pendant active", "Heure 23:00 → modifier à 14:00 en marche", "T-001 OK + schedule actif", "—", "_on_schedule_time_change reconfigure proprement", "Mineur"),
    T(130, "Robustesse", "REQ-ROB-15", "Onglet changé pendant scan", "Lancer scan, switch sur autre onglet", "T-001 OK", "input_nominal/", "Scan continue, progress affichée dans status bar globale", "Majeur"),
    T(131, "Robustesse", "REQ-ROB-16", "Window resize pendant scan", "Redimensionner pendant scan", "T-001 OK", "input_volumineux/", "Pas de freeze, layout adapté", "Mineur"),
    T(132, "Robustesse", "REQ-ROB-17", "Doublons quitter pendant scan", "Fermer app pendant scan doublons", "T-001 OK", "input_doublons/", "Cleanup, pas de processus orphelin", "Majeur"),
    T(133, "Robustesse", "REQ-ROB-18", "Tooltip après destruction widget", "Afficher tooltip puis détruire widget", "T-001 OK", "—", "Tooltip se ferme proprement, pas d'AttributeError", "Mineur"),
    T(134, "Robustesse", "REQ-ROB-19", "Logger encoding UTF-8", "Log message avec emojis", "T-001 OK", "—", "Fichier log sans corruption (utf-8 strict)", "Mineur"),
    T(135, "Robustesse", "REQ-ROB-20", "Plyer absent fallback", "Désinstaller plyer puis run", "T-001 OK", "—", "Toast remplacée par CTkToplevel auto-dismiss", "Mineur"),
]

# =============================================================================
# Régression (T-136 à T-150)
# =============================================================================
tests += [
    T(136, "Régression", "REQ-REG-01", "Burst detection vs ref", "F08 reproductibilité", "T-001 OK", "input_bursts/", "diff outputs_reels vs outputs_reference = 0", "Majeur"),
    T(137, "Régression", "REQ-REG-02", "Doublons DRY_RUN reproductible", "F22 même résultat 2 runs", "T-001 OK", "input_doublons/", "Mêmes groupes, mêmes décisions", "Majeur"),
    T(138, "Régression", "REQ-REG-03", "Invariant d'annulation", "Vérifier réussi+échoué+ignoré == total", "T-001 OK", "—", "Invariant arithmétique respecté", "Bloquant"),
    T(139, "Régression", "REQ-REG-04", "Scheduler trigger 1×/jour", "Trigger une fois, attendre 2 min même HH:MM", "T-001 OK", "—", "_last_run_date check : 1 seul trigger", "Majeur"),
    T(140, "Régression", "REQ-REG-05", "Tooltips après refonte", "Survoler 25 widgets clés (dont brand_examples_btn 💡 nouveau v2.3)", "T-001 OK", "—", "Tous affichent texte FR (registry has_tooltip) + clés filter_camera_make / brand_examples_btn présentes", "Mineur"),
    T(141, "Régression", "REQ-REG-06", "Cold-start non dégradé", "Compare cold-start v3 vs v4", "Build OK", "—", "Régression < +500 ms", "Mineur"),
    T(142, "Régression", "REQ-REG-07", "Index CSV columns stables", "Run organisation avec export CSV", "T-001 OK", "input_nominal/", "Mêmes colonnes que ref_T-142.csv", "Majeur"),
    T(143, "Régression", "REQ-REG-08", "Mode incrémental cache invariant", "F09 2 runs", "T-001 OK", "input_nominal/", ".photoorganizer_index.json grandit, pas de réécriture totale", "Mineur"),
    T(144, "Régression", "REQ-REG-09", "GPS extraction piexif fix", "Vérifier coords extraites de rationnel piexif", "T-001 OK", "input_gps_piexif/", "Coords non None (cf. fix _to_decimal)", "Majeur"),
    T(145, "Régression", "REQ-REG-10", "Drag-drop fonctionnel EXE", "DnD dans EXE PyInstaller", "Build OK", "input_nominal/", "tkdnd 2.9.4 actif, drop fonctionnel", "Majeur"),
    T(146, "Régression", "REQ-REG-11", "Icône taskbar Windows", "EXE lancé, vérif barre tâches", "Build OK", "—", "Icône assets/icons/icon.ico visible", "Mineur"),
    T(147, "Régression", "REQ-REG-12", "Pytest non slow vert", "pytest tests/ -m 'not slow'", "—", "—", "98/98 verts", "Bloquant"),
    T(148, "Régression", "REQ-REG-13", "Pytest slow vert", "pytest tests/ -m 'slow'", "—", "—", "Tests volume/stress verts", "Majeur"),
    T(149, "Régression", "REQ-REG-14", "Ruff clean", "ruff check src/", "—", "—", "All checks passed", "Mineur"),
    T(150, "Régression", "REQ-REG-15", "Audit visuel 12/12", "python tools/visual_audit.py", "—", "—", "12/12 invariants OK", "Majeur"),
]

# =============================================================================
# Refonte v2.3 + finition 2026-05-18 (T-151 à T-165)
# =============================================================================
# 4 onglets internes du panneau Organisation, bouton 💡 Exemples de marques,
# panneau intégré qui remplace toutes les anciennes fenêtres modales.
tests += [
    # --- IHM : onglets internes + 💡 + panneau intégré (T-151 à T-159) ---
    T(151, "IHM", "REQ-IHM-31", "Onglets internes Organisation (×4)",
      "Vérifier que le panneau Organisation contient un bandeau d'onglets interne avec 4 onglets exacts : 🔍 Filtrer / 🗂 Organiser / 🛠 Traiter / 🏷 Renommer",
      "T-001 OK", "—",
      "Les 4 onglets exacts sont présents dans l'ordre attendu",
      "Majeur"),
    T(152, "IHM", "REQ-IHM-32", "Onglet par défaut = Organiser",
      "À l'ouverture du panneau Organisation, l'onglet 'Organiser' est sélectionné",
      "T-001 OK", "—",
      "_main_tabview.get() == '🗂  Organiser'",
      "Mineur"),
    T(153, "IHM", "REQ-IHM-33", "Bouton 💡 Exemples de marques",
      "Le bouton 💡 est présent à côté du champ 'Limiter aux marques' dans l'onglet Organiser",
      "T-001 OK", "—",
      "brand_examples_btn existe, text='💡', width=40, height=32, dans _tab_organize",
      "Majeur"),
    T(154, "IHM", "REQ-IHM-34", "Tooltip brand_examples_btn",
      "Survoler 600 ms le bouton 💡",
      "T-001 OK", "—",
      "Tooltip FR affiché (clé 'brand_examples_btn' présente dans ORGANIZE)",
      "Mineur"),
    T(155, "IHM", "REQ-IHM-35", "API du panneau intégré",
      "Ouvrir un panneau intégré (titre, constructeur, boutons de pied) puis fermer",
      "T-001 OK", "—",
      "Panneau créé pendant l'affichage, masque les onglets internes ; la fermeture détruit le panneau et restaure les onglets",
      "Majeur"),
    T(156, "IHM", "REQ-IHM-36", "Aperçu à blanc = panneau intégré",
      "Cliquer le bouton « Aperçu » du panneau latéral droit",
      "T-001 OK + source/destination valides", "input_nominal/",
      "Panneau intégré « 👁 Aperçu (à blanc, sans modification disque) » affiché — AUCUNE nouvelle fenêtre créée (refonte 2026-05-18)",
      "Majeur"),
    T(157, "IHM", "REQ-IHM-37", "Organisation terminée = panneau intégré",
      "Lancer une organisation simple, attendre la fin",
      "T-001 OK", "input_nominal/",
      "Panneau intégré « ✅ Organisation terminée » avec bouton « 📂 Ouvrir destination » — AUCUNE nouvelle fenêtre",
      "Majeur"),
    T(158, "IHM", "REQ-IHM-38", "Sauvegarder un preset = panneau intégré + validation",
      "Cliquer 💾 Sauvegarder un preset, saisir un nom avec espace, valider",
      "T-001 OK", "—",
      "Panneau intégré « Sauvegarder un preset » affiche une erreur en ligne rouge (aucune boîte de dialogue système) ; nom valide → preset enregistré + panneau fermé",
      "Majeur"),
    T(159, "IHM", "REQ-IHM-39", "Liste des fichiers détectés = panneau intégré",
      "Cliquer le bouton 📋 (compteur de fichiers) avec une source valide",
      "T-001 OK", "input_nominal/",
      "Panneau intégré « 📋 N fichier(s) détecté(s) » liste jusqu'à 500 chemins (sinon « … et N de plus ») — AUCUNE nouvelle fenêtre",
      "Mineur"),

    # --- Paramètres : comportements du panneau 💡 (T-160 à T-164) ---
    T(160, "Paramètres", "REQ-PRM-21", "Panneau 💡 ajoute marque au champ",
      "Ouvrir panneau Exemples de marques, cliquer le bouton 'Sony'",
      "T-001 OK", "—",
      "filter_camera_make.get() == 'Sony' (ajout en CSV au champ)",
      "Majeur"),
    T(161, "Paramètres", "REQ-PRM-22", "Panneau 💡 dédupliquer",
      "Cliquer 'Sony' une 2e fois dans le panneau",
      "T-160 OK", "—",
      "filter_camera_make.get() == 'Sony' (pas de doublon 'Sony,Sony')",
      "Mineur"),
    T(162, "Paramètres", "REQ-PRM-23", "Panneau 💡 bouton 🗑 Vider",
      "Champ = 'Sony,Canon', cliquer '🗑 Vider' dans le panneau",
      "T-160 OK", "—",
      "filter_camera_make.get() == ''",
      "Mineur"),
    T(163, "Paramètres", "REQ-PRM-24", "Panneau 💡 marques détectées source",
      "Source contient photos EXIF Sony+Canon, ouvrir panneau 💡",
      "T-001 OK", "input_gps_piexif/",
      "Section '📷 Détectées dans la source' affiche les marques EXIF de l'échantillon (50 premiers fichiers)",
      "Mineur"),
    T(164, "Paramètres", "REQ-PRM-25", "COMMON_CAMERA_MAKES exposé",
      "Vérifier la constante OrganizeFrame.COMMON_CAMERA_MAKES",
      "T-001 OK", "—",
      "Tuple ≥ 10 marques, trié alphabétiquement, inclut Apple/Canon/Nikon/Sony",
      "Cosmétique"),

    # --- Régression : invariant 'pas de Toplevel' (T-165) ---
    T(165, "Régression", "REQ-REG-16", "Pas de Toplevel sur actions Organize",
      "Déclencher Aperçu / 📋 / 💡 / Sauvegarder un preset et compter les nouvelles fenêtres ouvertes",
      "T-001 OK", "—",
      "count_toplevels(app) invariant avant/après chaque action (mémo utilisateur : 'pas de nouvelles fenêtres dans Organisation')",
      "Majeur"),
]

# =============================================================================
# Refonte 2026-05-19 — bouton 💡 Exemples de filtres (T-166 à T-175)
# =============================================================================
# Panneau intégré ouvert depuis l'onglet Filtrer qui propose des valeurs
# standards pour aider l'utilisateur à remplir sans erreur les champs de
# pré-filtre (extensions, dimensions, mots-clés, orientation, note).
# La date et la taille en octets restent à saisir librement car propres
# à chaque utilisateur (consigne 2026-05-19).
tests += [
    T(166, "IHM", "REQ-IHM-40", "Bouton 💡 Exemples de filtres présent",
      "Vérifier la présence du bouton 💡 à côté du titre « 🔍 Filtres » dans l'onglet Filtrer",
      "T-001 OK", "—",
      "filter_examples_btn existe, libellé « 💡 », taille 40×32, situé dans le tab Filtrer",
      "Majeur"),
    T(167, "IHM", "REQ-IHM-41", "Tooltip Exemples de filtres",
      "Survoler 600 ms le bouton 💡 Exemples de filtres",
      "T-001 OK", "—",
      "Info-bulle FR affichée (clé « filter_examples_btn » présente dans ORGANIZE)",
      "Mineur"),
    T(168, "IHM", "REQ-IHM-42", "Panneau Exemples de filtres = intégré",
      "Cliquer le bouton 💡 de l'onglet Filtrer",
      "T-001 OK", "—",
      "Panneau intégré « 💡 Exemples de filtres (cliquer pour appliquer) » affiché — AUCUNE nouvelle fenêtre",
      "Majeur"),
    T(169, "Paramètres", "REQ-PRM-26", "Clic mot-clé ajoute au CSV",
      "Champ Mots-clés vide, ouvrir panneau, cliquer « vacances »",
      "T-001 OK", "—",
      "filter_keywords.get() == « vacances » ; un 2e clic ne crée pas de doublon",
      "Majeur"),
    T(170, "Paramètres", "REQ-PRM-27", "Clic extension cumule dans le CSV",
      "Champ Extensions vide, cliquer « jpg » puis « cr2 »",
      "T-001 OK", "—",
      "filter_extensions.get() == « jpg,cr2 » (CSV cumulatif déduplicaté)",
      "Majeur"),
    T(171, "Paramètres", "REQ-PRM-28", "Dimensions ↧ min / ↥ max",
      "Cliquer ↧ min sur la plus petite dimension, puis ↥ max sur la plus grande",
      "T-001 OK", "—",
      "filter_dim_min = première valeur (640×480) ; filter_dim_max = dernière (8K UHD)",
      "Majeur"),
    T(172, "Paramètres", "REQ-PRM-29", "Orientation applique la valeur",
      "Cliquer « Portrait » dans la section Orientation du panneau",
      "T-001 OK", "—",
      "filter_orientation.get() == « portrait »",
      "Mineur"),
    T(173, "Paramètres", "REQ-PRM-30", "Note minimale applique la valeur",
      "Cliquer le bouton 4 étoiles dans la section Note",
      "T-001 OK", "—",
      "filter_rating_min.get() == 4",
      "Mineur"),
    T(174, "Paramètres", "REQ-PRM-31", "Liste des constantes standards",
      "Vérifier la présence et la complétude des constantes de classe",
      "T-001 OK", "—",
      "COMMON_KEYWORDS ≥ 10 ; COMMON_EXTENSIONS_IMAGES contient jpg ; COMMON_EXTENSIONS_RAW contient cr2 ; COMMON_EXTENSIONS_VIDEOS contient mp4 ; COMMON_DIMENSIONS inclut Full HD ; COMMON_ORIENTATIONS == 4 (Toutes/Paysage/Portrait/Carré)",
      "Cosmétique"),
    T(175, "Régression", "REQ-REG-17", "Pas de Toplevel sur panneau Filtres",
      "Cliquer le bouton 💡 de l'onglet Filtrer et compter les nouvelles fenêtres",
      "T-001 OK", "—",
      "count_toplevels(app) invariant avant/après le clic (mémo utilisateur : « pas de nouvelles fenêtres dans Organisation »)",
      "Majeur"),
]

# =============================================================================
# Refonte 2026-05-19 — Quarantaine réversible (T-176 à T-189)
# =============================================================================
# Problème résolu : avant cette refonte, le mode TRASH du panneau Doublons
# envoyait les fichiers directement en corbeille système via send2trash
# (aller simple côté Python : aucune API de restauration). Conséquence :
# les boutons « ↩️ Annuler dernière » et « ↩️ Annuler tout » du panneau
# Historique ne pouvaient pas défaire une suppression de doublon.
# Solution : quarantaine interne via shutil.move + manifest JSON →
# rollback trivial. Le bouton « 🗑 Vider quarantaine » envoie ensuite
# en corbeille système quand l'utilisateur est sûr.
tests += [
    T(176, "Robustesse", "REQ-ROB-21", "Quarantaine : déplacement vers session_dir",
      "Appeler QuarantineManager.quarantine_file sur un fichier",
      "T-001 OK", "Fichier source existant",
      "Le fichier disparaît de la source et apparaît dans <root>/<session_id>/<hash>_<nom>",
      "Majeur"),
    T(177, "Robustesse", "REQ-ROB-22", "Quarantaine : manifest JSON persistant",
      "Vérifier <session_dir>/manifest.json après chaque opération",
      "T-176 OK", "—",
      "manifest.json existe, contient session_id + entries avec source/destination/timestamp",
      "Majeur"),
    T(178, "Robustesse", "REQ-ROB-23", "Quarantaine : pas de collision de noms",
      "Quarantiner 2 fichiers de même nom dans dossiers différents",
      "T-176 OK", "—",
      "Les 2 fichiers coexistent en quarantaine (préfixes de hash différents)",
      "Majeur"),
    T(179, "Cas limites", "REQ-LIM-26", "Quarantaine : fichier source absent",
      "Appeler quarantine_file sur un chemin inexistant",
      "T-001 OK", "—",
      "FileNotFoundError levée immédiatement avec message clair",
      "Mineur"),
    T(180, "Robustesse", "REQ-ROB-24", "Restauration : fichier remis à la source",
      "Quarantiner puis restore_entry du même fichier",
      "T-176 OK", "—",
      "Fichier réapparu à son chemin d'origine, entrée disparue du manifest",
      "Majeur"),
    T(181, "Cas limites", "REQ-LIM-27", "Restauration : dossier source supprimé",
      "Quarantiner, puis supprimer le dossier source, puis restore_entry",
      "T-180 OK", "—",
      "Le dossier source est recréé, le fichier est restauré dedans",
      "Majeur"),
    T(182, "Cas limites", "REQ-LIM-28", "Restauration : nom déjà repris à la source",
      "Quarantiner, recréer un fichier au même nom à la source, restaurer",
      "T-180 OK", "—",
      "Le restauré est suffixé « _restored1 », l'existant n'est pas écrasé",
      "Majeur"),
    T(183, "Sorties", "REQ-SOR-16", "Vider quarantaine → corbeille système",
      "Appeler empty_to_system_trash sur 3 fichiers en quarantaine",
      "T-176 OK", "send2trash dispo",
      "result['trashed'] == 3, manifest vidé, dossier session_dir nettoyé",
      "Majeur"),
    T(184, "Robustesse", "REQ-ROB-25", "Persistance : rechargement manifest",
      "Créer quarantaine, créer une nouvelle instance via load_session",
      "T-177 OK", "—",
      "load_session restitue toutes les entrées avec leurs métadonnées",
      "Majeur"),
    T(185, "Régression", "REQ-REG-18", "FileManager.record_operation enregistre 'trash'",
      "Appeler record_operation(operation_type='trash', ...) puis vérifier l'historique",
      "T-001 OK", "—",
      "L'opération apparaît dans get_operations_history() avec op_type 'trash'",
      "Majeur"),
    T(186, "Régression", "REQ-REG-19", "rollback_last sur 'trash' restaure le fichier",
      "Quarantiner + record_operation + rollback_last",
      "T-185 OK", "—",
      "Retourne True, fichier remis à la source, historique vidé",
      "Majeur"),
    T(187, "Régression", "REQ-REG-20", "rollback_all : trash récupéré, delete skipped",
      "Quarantine + record trash + delete réel + record delete, puis rollback_all",
      "T-185 OK", "—",
      "result['success']==1 (trash), result['skipped']==1 (delete définitif), failed==0",
      "Majeur"),
    T(188, "IHM", "REQ-IHM-43", "Bouton 🗑 Vider quarantaine présent",
      "Vérifier la présence du bouton dans le panneau bas de l'onglet Doublons",
      "T-001 OK", "—",
      "empty_quarantine_button existe, état initial = disabled (0 fichiers), libellé '🗑 Vider quarantaine (0)'",
      "Majeur"),
    T(189, "IHM", "REQ-IHM-44", "DuplicateManager expose l'API quarantaine",
      "Inspecter dm.empty_quarantine / quarantine_count / quarantine_size_bytes",
      "T-001 OK", "—",
      "Les 3 méthodes sont callables, instance vide retourne 0/0",
      "Cosmétique"),
]

# =============================================================================
# Sanity check : numérotation continue
# =============================================================================
ids = [t[0] for t in tests]
assert ids == [f"T-{i:03d}" for i in range(1, 190)], "Numérotation cassée !"
assert len(tests) == 189


# ---------------------------------------------------------------------------
# Génération du fichier xlsx
# ---------------------------------------------------------------------------
def main():
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.abspath(os.path.join(here, "..", "matrice_tests.xlsx"))

    wb = Workbook()

    # --- Feuille principale : Tests ---
    ws = wb.active
    ws.title = "Tests"

    # En-tête
    header_fill = PatternFill(start_color="1F6AA5", end_color="1F6AA5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Side(style="thin", color="666666")
    cell_border = Border(left=border, right=border, top=border, bottom=border)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = cell_border

    # Lignes de données
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, test in enumerate(tests, start=2):
        for col_idx, value in enumerate(test, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.alignment = body_align
            c.border = cell_border

    # Couleurs par catégorie
    cat_colors = {
        "IHM":          "E3F2FD",
        "Paramètres":   "F3E5F5",
        "Entrées":      "E8F5E9",
        "Sorties":      "FFF9C4",
        "Cas limites":  "FFE0B2",
        "Performance":  "FCE4EC",
        "Robustesse":   "FFCDD2",
        "Régression":   "D7CCC8",
    }
    for row_idx, test in enumerate(tests, start=2):
        cat = test[1]
        if cat in cat_colors:
            fill = PatternFill(start_color=cat_colors[cat], end_color=cat_colors[cat], fill_type="solid")
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Largeurs colonnes
    widths = {
        "A": 10, "B": 13, "C": 15, "D": 25, "E": 50, "F": 22, "G": 25,
        "H": 45, "I": 20, "J": 11, "K": 12, "L": 12, "M": 12, "N": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze première ligne
    ws.freeze_panes = "A2"

    # Validation Statut : OK / NOK / NA / À faire
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_statut = DataValidation(type="list",
                                formula1='"OK,NOK,NA,À faire"',
                                allow_blank=True)
    dv_statut.add(f"J2:J{len(tests) + 1}")
    ws.add_data_validation(dv_statut)

    dv_severite = DataValidation(type="list",
                                  formula1='"Bloquant,Majeur,Mineur,Cosmétique"',
                                  allow_blank=True)
    dv_severite.add(f"K2:K{len(tests) + 1}")
    ws.add_data_validation(dv_severite)

    # --- Feuille Synthèse ---
    ws_syn = wb.create_sheet("Synthèse")

    title_font = Font(bold=True, size=14, color="1F6AA5")
    ws_syn["A1"] = "Synthèse des tests — PhotoOrganizer v2.3-dev (refonte onglets internes + panneau intégré)"
    ws_syn["A1"].font = title_font
    ws_syn.merge_cells("A1:E1")

    # Tableau par catégorie
    ws_syn["A3"] = "Catégorie"
    ws_syn["B3"] = "Total"
    ws_syn["C3"] = "OK"
    ws_syn["D3"] = "NOK"
    ws_syn["E3"] = "NA"
    ws_syn["F3"] = "À faire"
    ws_syn["G3"] = "Taux OK (%)"
    for col_letter in "ABCDEFG":
        c = ws_syn[f"{col_letter}3"]
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border

    categories = ["IHM", "Paramètres", "Entrées", "Sorties", "Cas limites",
                   "Performance", "Robustesse", "Régression"]
    for i, cat in enumerate(categories, start=4):
        ws_syn[f"A{i}"] = cat
        ws_syn[f"B{i}"] = f'=COUNTIF(Tests!B:B,A{i})'
        ws_syn[f"C{i}"] = f'=COUNTIFS(Tests!B:B,A{i},Tests!J:J,"OK")'
        ws_syn[f"D{i}"] = f'=COUNTIFS(Tests!B:B,A{i},Tests!J:J,"NOK")'
        ws_syn[f"E{i}"] = f'=COUNTIFS(Tests!B:B,A{i},Tests!J:J,"NA")'
        ws_syn[f"F{i}"] = f'=COUNTIFS(Tests!B:B,A{i},Tests!J:J,"À faire")'
        ws_syn[f"G{i}"] = f'=IF(B{i}=0,"-",ROUND(C{i}/B{i}*100,1))'
        for col_letter in "ABCDEFG":
            ws_syn[f"{col_letter}{i}"].border = cell_border

    # Ligne TOTAL
    total_row = 4 + len(categories)
    ws_syn[f"A{total_row}"] = "TOTAL"
    ws_syn[f"A{total_row}"].font = Font(bold=True)
    ws_syn[f"B{total_row}"] = f"=SUM(B4:B{total_row-1})"
    ws_syn[f"C{total_row}"] = f"=SUM(C4:C{total_row-1})"
    ws_syn[f"D{total_row}"] = f"=SUM(D4:D{total_row-1})"
    ws_syn[f"E{total_row}"] = f"=SUM(E4:E{total_row-1})"
    ws_syn[f"F{total_row}"] = f"=SUM(F4:F{total_row-1})"
    ws_syn[f"G{total_row}"] = f"=IF(B{total_row}=0,\"-\",ROUND(C{total_row}/B{total_row}*100,1))"
    for col_letter in "ABCDEFG":
        c = ws_syn[f"{col_letter}{total_row}"]
        c.font = Font(bold=True)
        c.border = cell_border
        c.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Tableau par sévérité
    sev_row = total_row + 3
    ws_syn[f"A{sev_row}"] = "Sévérité"
    ws_syn[f"B{sev_row}"] = "Total"
    ws_syn[f"C{sev_row}"] = "OK"
    ws_syn[f"D{sev_row}"] = "NOK"
    for col_letter in "ABCD":
        c = ws_syn[f"{col_letter}{sev_row}"]
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border

    for i, sev in enumerate(["Bloquant", "Majeur", "Mineur", "Cosmétique"], start=sev_row + 1):
        ws_syn[f"A{i}"] = sev
        ws_syn[f"B{i}"] = f'=COUNTIF(Tests!K:K,A{i})'
        ws_syn[f"C{i}"] = f'=COUNTIFS(Tests!K:K,A{i},Tests!J:J,"OK")'
        ws_syn[f"D{i}"] = f'=COUNTIFS(Tests!K:K,A{i},Tests!J:J,"NOK")'
        for col_letter in "ABCD":
            ws_syn[f"{col_letter}{i}"].border = cell_border

    ws_syn.column_dimensions["A"].width = 18
    for col_letter in "BCDEFG":
        ws_syn.column_dimensions[col_letter].width = 12

    # --- Feuille Exigences ---
    ws_req = wb.create_sheet("Exigences")
    ws_req["A1"] = "Référentiel d'exigences fonctionnelles"
    ws_req["A1"].font = title_font
    ws_req.merge_cells("A1:C1")

    ws_req["A3"] = "ID Exigence"
    ws_req["B3"] = "Catégorie"
    ws_req["C3"] = "Tests couvrants"
    for col_letter in "ABC":
        c = ws_req[f"{col_letter}3"]
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border

    # Liste unique des exigences avec leurs tests
    from collections import defaultdict
    req_to_tests = defaultdict(list)
    for t in tests:
        req_to_tests[t[2]].append(t[0])

    for i, (req, t_ids) in enumerate(sorted(req_to_tests.items()), start=4):
        cat_pref = req.split("-")[1] if "-" in req else ""
        ws_req[f"A{i}"] = req
        ws_req[f"B{i}"] = {"IHM": "IHM", "PRM": "Paramètres", "ENT": "Entrées",
                            "SOR": "Sorties", "LIM": "Cas limites",
                            "PRF": "Performance", "ROB": "Robustesse",
                            "REG": "Régression"}.get(cat_pref, "-")
        ws_req[f"C{i}"] = ", ".join(t_ids)
        for col_letter in "ABC":
            ws_req[f"{col_letter}{i}"].border = cell_border
            ws_req[f"{col_letter}{i}"].alignment = body_align

    ws_req.column_dimensions["A"].width = 15
    ws_req.column_dimensions["B"].width = 15
    ws_req.column_dimensions["C"].width = 40

    # --- Feuille Légende ---
    ws_leg = wb.create_sheet("Légende")
    ws_leg["A1"] = "Légende de la matrice"
    ws_leg["A1"].font = title_font

    ws_leg["A3"] = "Statut"
    ws_leg["B3"] = "Description"
    ws_leg["A4"] = "OK"; ws_leg["B4"] = "Test exécuté et conforme au résultat attendu"
    ws_leg["A5"] = "NOK"; ws_leg["B5"] = "Test exécuté, résultat non conforme"
    ws_leg["A6"] = "NA"; ws_leg["B6"] = "Non applicable (feature désactivée, dépendance absente, etc.)"
    ws_leg["A7"] = "À faire"; ws_leg["B7"] = "Test pas encore exécuté (état initial)"

    ws_leg["A9"] = "Sévérité"
    ws_leg["B9"] = "Description"
    ws_leg["A10"] = "Bloquant"; ws_leg["B10"] = "Empêche l'utilisation de la fonctionnalité ou de l'outil entier"
    ws_leg["A11"] = "Majeur"; ws_leg["B11"] = "Dysfonctionnement important mais contournement possible"
    ws_leg["A12"] = "Mineur"; ws_leg["B12"] = "Anomalie sans impact significatif sur l'usage"
    ws_leg["A13"] = "Cosmétique"; ws_leg["B13"] = "Imperfection visuelle ou rédactionnelle"

    ws_leg["A15"] = "Catégorie"
    ws_leg["B15"] = "Périmètre"
    ws_leg["A16"] = "IHM"; ws_leg["B16"] = "Comportement visuel et interactions widgets"
    ws_leg["A17"] = "Paramètres"; ws_leg["B17"] = "Configuration utilisateur (thème, planification, cache, …)"
    ws_leg["A18"] = "Entrées"; ws_leg["B18"] = "Types et formes des données en entrée"
    ws_leg["A19"] = "Sorties"; ws_leg["B19"] = "Arborescences/rapports/index produits"
    ws_leg["A20"] = "Cas limites"; ws_leg["B20"] = "Bornes, scénarios exotiques, gestion d'erreurs"
    ws_leg["A21"] = "Performance"; ws_leg["B21"] = "Temps de réponse, mémoire, scalabilité"
    ws_leg["A22"] = "Robustesse"; ws_leg["B22"] = "Stabilité sous spam/concurrence/conditions dégradées"
    ws_leg["A23"] = "Régression"; ws_leg["B23"] = "Non-régression sur features post-refonte"

    for row in [3, 9, 15]:
        for col_letter in "AB":
            c = ws_leg[f"{col_letter}{row}"]
            c.font = header_font
            c.fill = header_fill

    ws_leg.column_dimensions["A"].width = 15
    ws_leg.column_dimensions["B"].width = 60

    # Sauvegarde
    wb.save(out)
    print(f"OK : {out}")
    print(f"  - {len(tests)} tests générés (T-001 à T-{len(tests):03d})")
    print("  - 4 feuilles : Tests, Synthèse, Exigences, Légende")
    print(f"  - 8 catégories : {', '.join(categories)}")
    print(f"  - {len(set(t[2] for t in tests))} exigences référencées")


if __name__ == "__main__":
    main()
